import sqlite3
import asyncio
import logging
import os
import re
from datetime import datetime, timedelta
from typing import List

from aiogram import Bot, Dispatcher, types
from aiogram.types import (ParseMode, BotCommand, ReplyKeyboardMarkup, 
                          KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton)
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils import executor
from aiohttp import web

import csv
import io
from aiogram.types import InputFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from aiogram.utils import exceptions
from aiogram.types import ChatMemberUpdated, ChatType

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Конфигурация
API_TOKEN = os.getenv('apibotkey')
DB_PATH = "/bd1/tasks.db"

# Список разрешенных пользователей
ALLOWED_USERS: List[str] = []  

def update_allowed_users(conn):
    global ALLOWED_USERS
    cursor = conn.cursor()
    cursor.execute('SELECT CAST(tg_user_id as INT) FROM users')
    ALLOWED_USERS = [row[0] for row in cursor.fetchall()]

# Список модераторов
MODERATOR_USERS: List[str] = []  

def update_moderator_users(conn):
    global MODERATOR_USERS
    cursor = conn.cursor()
    cursor.execute("""SELECT CAST(tg_user_id as INT) FROM users WHERE is_moderator = 'moderator' """)
    MODERATOR_USERS = [row[0] for row in cursor.fetchall()]

# ID администратора (может удалять задачи)
ADMIN_ID = int(os.getenv('admin'))

# Инициализация бота и диспетчера
bot = Bot(token=API_TOKEN, parse_mode=ParseMode.HTML)
dp = Dispatcher(bot, storage=MemoryStorage())

# Инициализация базы данных
def init_db():
    try:
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        cursor = conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS tasks (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        creator_id TEXT,
                        user_id TEXT,
                        chat_id INTEGER,
                        task_text TEXT,
                        status TEXT DEFAULT 'новая',
                        deadline TEXT)
                        ''')
        conn.commit()

        cursor.execute('''CREATE TABLE IF NOT EXISTS users (
                        tg_user_id TEXT PRIMARY KEY,
                        name TEXT,
                        username TEXT,
                        is_moderator TEXT)
                        ''')
        conn.commit()

        cursor.execute('''CREATE TABLE IF NOT EXISTS tasks_log (
                        id INTEGER,
                        creator_id TEXT,
                        user_id TEXT,
                        chat_id INTEGER,
                        task_text TEXT,
                        status TEXT,
                        deadline TEXT,
                        priority TEXT,
                        id_log INTEGER PRIMARY KEY AUTOINCREMENT)
                        ''')

        # Индексы при инициализации БД
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_tasks_user_id ON tasks(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_tasks_chat_id ON tasks(chat_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_tasks_creator_id ON tasks(creator_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_tasks_deadline ON tasks(deadline)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_tasks_log_id ON tasks_log(id)')
        
        conn.commit()
      
        return conn
    except sqlite3.Error as e:
        logger.error(f"Ошибка при инициализации БД: {e}")
        raise

# Для фоновых задач
def create_db_connection():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

conn = init_db()
update_allowed_users(conn)
update_moderator_users(conn)

background_conn = create_db_connection()

# ======================
# КЛАВИАТУРЫ И ИНТЕРФЕЙС
# ======================

# Главное меню
menu_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
menu_keyboard.add(
    KeyboardButton("➕ Новая задача"),
    KeyboardButton("⚡ Быстрая задача"),
    KeyboardButton("🔄 Изменить статус"),
    KeyboardButton("✏️ Изменить задачу"),
    KeyboardButton("👤 Изменить исполнителя"),
    KeyboardButton("⏳ Изменить срок"),
    KeyboardButton("📋 Список задач"),
    KeyboardButton("📋 Список (по сроку)"),
    KeyboardButton("📤 Экспорт задач"),
    KeyboardButton("📤 Экспорт (с исполненными)"),
    KeyboardButton("⛔ Отмена")
)

# Клавиатура для групповых чатов
group_menu_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
group_menu_keyboard.add(
    KeyboardButton("⚡ Быстрая задача"),
    KeyboardButton("📤 Экспорт задач"),
    KeyboardButton("⛔ Отмена")
)

# Клавиатура выбора даты
def get_deadline_keyboard(with_none_option=False):
    today = datetime.today()
    dates = {
        "Сегодня": today.strftime("%Y-%m-%d"),
        "Завтра": (today + timedelta(days=1)).strftime("%Y-%m-%d"),
        "Послезавтра": (today + timedelta(days=2)).strftime("%Y-%m-%d"),
    }

    # Дополнительные кнопки
    if with_none_option:
        dates["❌ Без срока"] = "set_deadline_none"
    dates["Свой срок"] = "set_deadline_custom"

    keyboard = InlineKeyboardMarkup(row_width=3)  # 3 кнопки в ряду
    
    # Добавляем кнопки парами
    buttons = []
    for label, date in dates.items():
        if label == "❌ Без срока" or label == "Свой срок":
            buttons.append(InlineKeyboardButton(label, callback_data=date))  # для этих кнопок callback_data = date
        else:
            buttons.append(InlineKeyboardButton(label, callback_data=f"set_deadline_{date}"))  # для остальных - в формате f"set_deadline_{date}"

    # Распределяем кнопки по 3 в ряд
    for i in range(0, len(buttons), 3):
        row = buttons[i:i+3]
        keyboard.row(*row)

    return keyboard

# ======================
# ОБРАБОТЧИКИ КОМАНД
# ======================

# Установка команд с подсказками
async def set_bot_commands(bot: Bot):
    commands = [
        BotCommand(command="/newtask", description="Создать задачу"),
        BotCommand(command="/quicktask", description="Быстрая задача"),
        BotCommand(command="/setstatus", description="Изменить статус"),
        BotCommand(command="/settext", description="Изменить задачу"),
        BotCommand(command="/setexecutor", description="Изменить исполнителя"),
        BotCommand(command="/setdeadline", description="Изменить срок"),
        BotCommand(command="/listtasks", description="Список задач"),
        BotCommand(command="/listtasksdate", description="Список (по сроку)"),
        BotCommand(command="/export", description="Экспорт в CSV"),
        BotCommand(command="/export2", description="Экспорт (с исполненными)"),
        BotCommand(command="/start", description="Старт бота"),
        BotCommand(command="/cancel", description="Отмена текущего действия"),
        BotCommand(command="/myid", description="Узнать свой ID"),
        BotCommand(command="/export3", description="Полный экспорт (админ)"),
        BotCommand(command="/deletetask", description="Удалить задачу (админ)"),
        BotCommand(command="/export4", description="Список пользователей (админ)"),
        BotCommand(command="/adduser", description="Добавить пользователя (админ)"),
        BotCommand(command="/removeuser", description="Удалить пользователя (админ)")
    ]
    await bot.set_my_commands(commands)

@dp.message_handler(commands=["start"])
async def start_command(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS and message.from_user.id != ADMIN_ID:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return

    if message.chat.type == "private":
        await bot.send_message(chat_id=message.chat.id, text=
            "Выберите команду:",
            reply_markup=menu_keyboard
        )
    else:
        await bot.send_message(chat_id=message.chat.id, text=
            "Выберите команду:",
            reply_markup=group_menu_keyboard
        )

# Команды вызывают те же функции, что и кнопки
@dp.message_handler(commands=["newtask"])
async def cmd_new_task(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    await new_task_start(message)  # Тот же обработчик, что и для кнопки "➕ Новая задача"

@dp.message_handler(commands=["quicktask"])
async def cmd_quick_task(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    await quick_task_start(message)  # Тот же обработчик, что и для кнопки "⚡ Быстрая задача"

@dp.message_handler(commands=["setstatus"])
async def cmd_set_status(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Менять статус можно только в ЛС")
        return
    await status_select_task(message)  # Аналогично кнопке "🔄 Изменить статус"

@dp.message_handler(commands=["settext"])
async def cmd_set_status(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Менять задачу можно только в ЛС")
        return
    await text_edit_start(message)  # Аналогично кнопке "✏️ Изменить задачу"

@dp.message_handler(commands=["setexecutor"])
async def cmd_set_executor(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Менять исполнителя можно только в ЛС")
        return
    await executor_select_task(message)  # Аналогично кнопке "👤 Изменить исполнителя"

@dp.message_handler(commands=["setdeadline"])
async def cmd_set_deadline(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Менять срок можно только в ЛС")
        return
    await deadline_select_task(message)  # Аналогично кнопке "⏳ Изменить срок"

@dp.message_handler(commands=["listtasks"])
async def cmd_list_tasks(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Выводить список можно только в ЛС")
        return  
    await list_tasks(message)  # Аналогично кнопке "📋 Список задач"

@dp.message_handler(commands=["listtasksdate"])
async def cmd_list_tasks_date(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Выводить список можно только в ЛС")
        return  
    await list_tasks_by_deadline(message)  # Аналогично кнопке "📋 Список (по сроку)"

@dp.message_handler(commands=["export"])
async def cmd_export_tasks(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    await export_tasks_to_csv(message)  # Аналогично кнопке "📤 Экспорт задач"

@dp.message_handler(commands=["export2"])
async def cmd_export_tasks(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    await export_tasks_to_csv2(message)  # Аналогично кнопке "📤 Экспорт (с исполненными)"

@dp.message_handler(commands=["cancel"])
async def cmd_cancel(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    await cancel_handler(message)  # Тот же обработчик, что и для кнопки "⛔ Отмена"

# ======================
# СОСТОЯНИЯ БОТА
# ======================

class TaskCreation(StatesGroup):
    waiting_for_title = State()
    waiting_for_executor = State()
    waiting_for_deadline = State()

class TaskDeletion(StatesGroup):
    waiting_for_task_selection = State()
    waiting_for_confirmation = State()
    waiting_for_manual_id = State()

@dp.message_handler(lambda message: message.text == "⛔ Отмена", state='*')
async def cancel_handler(message: types.Message, state: FSMContext):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
      
    current_state = await state.get_state()
    if current_state:
        await state.finish()
    
    if message.chat.type != "private":
        await message.reply("Действие отменено. Возвращаемся к стартовому меню.", reply_markup=group_menu_keyboard)
    else:
        await message.reply("Действие отменено. Возвращаемся к стартовому меню.", reply_markup=menu_keyboard)

def format_date(date_str):
    try:
        # Пытаемся распарсить срок как дату со временем
        dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
    except Exception:
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            return date_str

    # Если время равно 00:00, выводим только дату
    if dt.hour == 0 and dt.minute == 0:
        return dt.strftime("%d.%m.%Y")
    else:
        # Выводим дату с временем в скобках, например "10.05.2025 (12:30)"
        return f"{dt.strftime('%d.%m.%Y')} ({dt.strftime('%H:%M')})"

# ======================
# СОЗДАНИЕ ЗАДАЧ
# ======================

@dp.message_handler(lambda message: message.text == "➕ Новая задача")
async def new_task_start(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return

    if message.chat.type != "private":
      await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
      return

    """Начало создания задачи через кнопку меню"""
    await bot.send_message(chat_id=message.from_user.id, text="📌 Введите название задачи:")
    await TaskCreation.waiting_for_title.set()

@dp.message_handler(state=TaskCreation.waiting_for_title)
async def process_title(message: types.Message, state: FSMContext):
    # Получаем список исполнителей из БД
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE status <> 'удалено' LIMIT 20")
    executors = [executor[0] for executor in cursor.fetchall() if executor[0]]

    # Создаём inline-клавиатуру (замена ReplyKeyboardMarkup)
    keyboard = InlineKeyboardMarkup(row_width=2)
    for i in range(0, len(executors), 2):
        row_buttons = [InlineKeyboardButton(f"👤 {name}", callback_data=f"executor_select|{name}")
                       for name in executors[i:i+2]]
        keyboard.row(*row_buttons)
    # Кнопка для ручного ввода
    keyboard.add(InlineKeyboardButton("✏️ Ввести @username вручную", callback_data="executor_select|manual"))

    await bot.send_message(
        chat_id=message.chat.id,
        text="👤 Выберите исполнителя или введите @username вручную:",
        reply_markup=keyboard
    )
    await state.update_data(title=message.text)
    await TaskCreation.waiting_for_executor.set()

@dp.callback_query_handler(lambda c: c.data.startswith("executor_select|"), state=TaskCreation.waiting_for_executor)
async def process_executor_callback(callback_query: types.CallbackQuery, state: FSMContext):
    executor = callback_query.data.split("|", 1)[1]
    if executor == "manual":
        await bot.answer_callback_query(callback_query.id, text="✏️ Введите @username вручную")
        # Дальше можно оставить ожидание текстового ввода (обработчик ниже уже существует)
        return
    # Сохраняем выбранного исполнителя в состоянии
    await state.update_data(executor=executor)
    # Убираем inline-клавиатуру, редактируя сообщение
    await bot.edit_message_reply_markup(chat_id=callback_query.message.chat.id,
                                        message_id=callback_query.message.message_id,
                                        reply_markup=None)
    # Переходим к следующему шагу: выбор срока
    await bot.send_message(
        chat_id=callback_query.from_user.id,
        text="⏳ Выберите срок или введите свой:",
        reply_markup=get_deadline_keyboard(with_none_option=True)
    )
    await TaskCreation.waiting_for_deadline.set()

@dp.message_handler(state=TaskCreation.waiting_for_executor)
async def process_executor(message: types.Message, state: FSMContext):
    """Обработка исполнителя задачи"""
    executor = message.text.strip()
    await state.update_data(executor=executor)
    
    # Убираем клавиатуру после выбора
    remove_kb = types.ReplyKeyboardRemove()
    await bot.send_message(
        chat_id=message.chat.id,
        text="⏳ Выберите срок или введите свой:",
        reply_markup=get_deadline_keyboard(with_none_option=True)
    )
    await TaskCreation.waiting_for_deadline.set()

@dp.callback_query_handler(lambda c: c.data.startswith("set_deadline_"), state=TaskCreation.waiting_for_deadline)
async def process_deadline(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка выбора дедлайна"""
    if callback_query.data == "set_deadline_custom":
        # Сохраняем callback_query в состоянии
        await state.update_data(callback_query=callback_query)
        await bot.send_message(chat_id=callback_query.from_user.id, text="⏳ Введите срок в формате DD.MM.YYYY:")
        return
    elif callback_query.data == "set_deadline_none":
        await save_task(callback_query, state, deadline=None)
    else:
        deadline = callback_query.data.split("_")[2]
        await save_task(callback_query, state, deadline)

@dp.message_handler(state=TaskCreation.waiting_for_deadline)
async def process_custom_deadline(message: types.Message, state: FSMContext):
    """Обработка ввода собственного срока"""
    try:
        # Попытка парсинга по первому шаблону
        try:
            dt = datetime.strptime(message.text.strip(), "%Y-%m-%d %H:%M")
        except ValueError:
            try:
                dt = datetime.strptime(message.text.strip(), "%d.%m.%Y %H:%M")
            except ValueError:
                try:
                    dt = datetime.strptime(message.text.strip(), "%d.%m.%y %H:%M")
                except ValueError:
                    try:
                        dt = datetime.strptime(message.text.strip(), "%Y-%m-%d")
                    except ValueError:
                        # Если не получилось, пробуем другой формат
                        try:
                            dt = datetime.strptime(message.text.strip(), "%d.%m.%Y")
                        except ValueError:
                            dt = datetime.strptime(message.text.strip(), "%d.%m.%y")
        new_deadline = dt.strftime("%Y-%m-%d %H:%M")
        
        # Получаем сохраненный callback_query из состояния (если он есть)
        user_data = await state.get_data()
        callback_query = user_data.get('callback_query')
        
        if callback_query:
            await save_task(callback_query, state, new_deadline)
        else:
            await save_task(message, state, new_deadline)
            
    except ValueError:
        # Определяем клавиатуру в зависимости от типа чата
        reply_markup = menu_keyboard if message.chat.type == "private" else group_menu_keyboard
        await bot.send_message(chat_id=message.chat.id, text="⚠ Ошибка! Введите дату в формате  или DD.MM.YYYY", reply_markup=reply_markup)
        await state.finish()

async def save_task(message_obj, state: FSMContext, deadline: str):
    """Сохранение задачи в БД"""
    user_data = await state.get_data()
    task_text = user_data['title']
    executor = user_data['executor']

    try:
        # Получаем chat_id и тип чата
        if isinstance(message_obj, types.CallbackQuery):
            chat_id = message_obj.from_user.id
            chat_id2 = message_obj.message.chat.id
            chat_type = message_obj.message.chat.type
            message_to_reply = message_obj.message
        else:  # Это обычное сообщение (types.Message)
            chat_id = message_obj.from_user.id
            chat_id2 = message_obj.message.chat.id
            chat_type = message_obj.chat.type
            message_to_reply = message_obj

        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO tasks (user_id, chat_id, task_text, deadline, creator_id) VALUES (?, ?, ?, ?, ?)",
            (executor, chat_id, task_text, deadline, chat_id)
        )
        conn.commit()

        cursor.execute("SELECT tg_user_id FROM users WHERE username=?",(executor,))
        username = cursor.fetchone()

        cursor.execute("SELECT username FROM users WHERE tg_user_id=?",(chat_id,))
        creator = cursor.fetchone()

        response = (
            f"📌 <b>{task_text}</b>\n"
            f"👤 {executor} "
        )

        response2 = (
            f"🔔 Вам назначена новая задача от {creator[0]}:\n\n"
            f"📌 <b>{task_text}</b>\n"
        )
        if deadline:
            response += f"⏳ {format_date(deadline)}"
            response2 += f"⏳ {format_date(deadline)}"
        else:
            response += "⏳ Без срока"
            response2 += "⏳ Без срока"
            
        # Определяем клавиатуру в зависимости от типа чата
        reply_markup = menu_keyboard if chat_type == "private" else group_menu_keyboard
        
        # Отправляем сообщение с клавиатурой
        await bot.send_message(
            chat_id=chat_id2,
            text=response,
            parse_mode=ParseMode.HTML,
            reply_markup=reply_markup
        )
        if username is not None and username[0] is not None and username[0] != str(chat_id):
            await bot.send_message(
                chat_id=username[0],
                text=response2,
                parse_mode=ParseMode.HTML
            )
  
    except sqlite3.Error as e:
        logger.error(f"Ошибка БД при сохранении задачи: {e}")
        reply_target = message_obj.message if isinstance(message_obj, types.CallbackQuery) else message_obj
        await reply_target.reply(f"⚠ Ошибка при сохранении задачи: {str(e)}")
    finally:
        await state.finish()

# ======================
# СОЗДАНИЕ ЗАДАЧИ ИЗ ОДНОГО СООБЩЕНИЯ
# ======================

def parse_deadline(deadline_str: str) -> str:
    """Преобразует текстовое представление даты в формат YYYY-MM-DD"""
    today = datetime.today()
    weekday_map = {
        'пн': 0, 'пон': 0, 'понедельник': 0,
        'вт': 1, 'вто': 1, 'вторник': 1,
        'ср': 2, 'сре': 2, 'среда': 2,
        'чт': 3, 'чет': 3, 'четверг': 3,
        'пт': 4, 'пят': 4, 'пятница': 4,
        'сб': 5, 'суб': 5, 'суббота': 5,
        'вс': 6, 'вос': 6, 'воскресенье': 6
    }
    
    lower_str = deadline_str.lower()
    
    if lower_str == 'сегодня':
        return today.strftime("%Y-%m-%d")
    
    if lower_str == 'завтра':
        return (today + timedelta(days=1)).strftime("%Y-%m-%d")
    
    if lower_str in weekday_map:
        target_weekday = weekday_map[lower_str]
        current_weekday = today.weekday()
        
        days_ahead = target_weekday - current_weekday
        if days_ahead <= 0:
            days_ahead += 7
            
        return (today + timedelta(days=days_ahead)).strftime("%Y-%m-%d")

    # Попытка парсинга по первому шаблону
    try:
        dt = datetime.strptime(deadline_str, "%Y-%m-%d %H:%M")
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        try:
            dt = datetime.strptime(deadline_str, "%d.%m.%Y %H:%M")
            return dt.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            try:
                dt = datetime.strptime(deadline_str, "%d.%m.%y %H:%M")
                return dt.strftime("%Y-%m-%d %H:%M")
            except ValueError:
                try:
                    dt = datetime.strptime(deadline_str, "%Y-%m-%d")
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    # Если не получилось, пробуем другой формат
                    try:
                        dt = datetime.strptime(deadline_str, "%d.%m.%Y")
                        return dt.strftime("%Y-%m-%d")
                    except ValueError:
                        try:
                            dt = datetime.strptime(deadline_str, "%d.%m.%y")
                            return dt.strftime("%Y-%m-%d")
                        except ValueError:
                            raise ValueError("Неверный формат даты. Используйте DD.MM.YYYY")

class QuickTaskCreation(StatesGroup):
    waiting_for_full_data = State()

@dp.message_handler(lambda message: message.text == "⚡ Быстрая задача")
async def quick_task_start(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return
    """Начало быстрого создания задачи"""
    await bot.send_message(chat_id=message.from_user.id, text=
        "📝 Введите данные в формате:\n"
        "текст задачи @исполнитель //срок"
    )
    await QuickTaskCreation.waiting_for_full_data.set()

@dp.message_handler(state=QuickTaskCreation.waiting_for_full_data,
                    content_types=types.ContentType.ANY)
async def process_quick_task(message: types.Message, state: FSMContext):
    """Обработка быстрого создания задачи"""
    try:
        text = message.text if message.text else (message.caption if message.caption else "")
        
        # Парсим данные с помощью регулярных выражений
        if text and text.startswith('@'):
            # Захватить всё до появления '//' или до конца строки.
            task_match = re.search(r'^(.*?)(?=//|$)', text)
        else:
            # Захватывать всё от начала до пробела перед @, если он есть, или до конца строки.
            task_match = re.search(r'^(.*?)(\s@|$)', text)
        executor_match = re.search(r'(@[^\s]+)', text)
        deadline_match = re.search(r'//\s*(.+)', text)
        deadline_raw = deadline_match.group(1) if deadline_match else None

        task_text = task_match.group(1).strip() if task_match else None
        executor = executor_match.group(0).strip() if executor_match else None
        deadline = deadline_match.group(1) if deadline_match else None

        # Валидация обязательного поля
        if not task_text:
            raise ValueError("Не указан текст задачи")

        # Проверка формата даты
        deadline_raw = deadline_match.group(1) if deadline_match else None
        deadline = None
        if deadline_raw:
            try:
                deadline = parse_deadline(deadline_raw)
            except ValueError as e:
                raise ValueError(f"Ошибка в сроке: {str(e)}")

        # Сохранение в БД
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO tasks (user_id, chat_id, task_text, deadline, creator_id) VALUES (?, ?, ?, ?, ?)",
            (executor, message.from_user.id, task_text, deadline, message.from_user.id)
        )
        conn.commit()

        cursor.execute("SELECT tg_user_id FROM users WHERE username=?",(executor,))
        username = cursor.fetchone()

        cursor.execute("SELECT username FROM users WHERE tg_user_id=?",(message.from_user.id,))
        creator = cursor.fetchone()

        response = (
            f"📌 <b>{task_text}</b>\n"
            f"👤 {executor if executor else 'не указан'} ⏳ {format_date(deadline) if deadline else 'не указан'}"
        )

        response2 = (
            f"🔔 Вам назначена новая задача от {creator[0]}:\n\n"
            f"📌 <b>{task_text}</b>\n"
            f"⏳ {format_date(deadline) if deadline else 'не указан'}"
        )
          
        await bot.send_message(chat_id=message.from_user.id, text=response)

        if username is not None and username[0] is not None and username[0] != str(message.from_user.id):
          await bot.send_message(
              chat_id=username[0],
              text=response2
          )

    except ValueError as e:
        await bot.send_message(chat_id=message.from_user.id,text=f"⚠ Ошибка: {str(e)}")
    except sqlite3.Error as e:
        logger.error(f"Ошибка БД: {e}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Ошибка при сохранении задачи")
    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Неверный формат данных/У исполнителя нет диалога с ботом")
    finally:
        await state.finish()

# ======================
# ИЗМЕНЕНИЕ СТАТУСА
# ======================

class StatusUpdate(StatesGroup):
    waiting_for_executor = State()
    waiting_for_task_selection = State()
    waiting_for_status_choice = State()

@dp.message_handler(lambda message: message.text == "🔄 Изменить статус")
async def status_select_task(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return 

    if message.chat.type != "private":
      await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
      return

    """Показ списка задач для изменения статуса"""
    
    # Сначала получаем список уникальных исполнителей
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT user_id FROM tasks 
        WHERE status NOT IN ('удалено', 'исполнено')
        LIMIT 20
    """)
    
    executors = cursor.fetchall()
    
    if not executors:
        await message.reply("❌ Нет задач для изменения статуса")
        return

    keyboard = InlineKeyboardMarkup(row_width=2)  # Устанавливаем количество кнопок в ряду
    
    # Разбиваем исполнителей на группы по 2
    for i in range(0, len(executors), 2):
        row = executors[i:i+2]  # Берем группу из 2 элементов
        row_buttons = [
            InlineKeyboardButton(
                f"👤 {executor[0] if executor[0] else 'Без исполнителя'}",
                callback_data=f"executor_for_status|{executor[0]}"
            ) for executor in row
        ]
        keyboard.add(*row_buttons)  # Добавляем группу кнопок в клавиатуру
    
    # Добавляем кнопку для ввода ID вручную
    keyboard.add(InlineKeyboardButton("✏️ Ввести ID задачи вручную", callback_data="status_manual_id"))
    
    await message.reply("Выберите исполнителя для фильтрации задач:", reply_markup=keyboard)
    await StatusUpdate.waiting_for_executor.set()

@dp.callback_query_handler(lambda c: c.data.startswith("executor_for_status|"), state=StatusUpdate.waiting_for_executor)
async def process_executor_selection(callback_query: types.CallbackQuery, state: FSMContext):
    executor = callback_query.data.split("|")[-1]
    await state.update_data(executor=executor)
    await show_filtered_tasks(callback_query.message, executor)
    await StatusUpdate.waiting_for_task_selection.set()

async def show_filtered_tasks(message_obj, executor):
    """Показать задачи выбранного исполнителя"""
    try:
        cursor = conn.cursor()
        if executor.lower() == "none":  # Проверяем, ищем ли задачи без исполнителя
            cursor.execute("""
                SELECT id, task_text, status 
                FROM tasks
                WHERE user_id IS NULL AND status NOT IN ('удалено', 'исполнено')
                ORDER BY id DESC 
                LIMIT 20
            """)
        else:
            cursor.execute("""
                SELECT id, task_text, status 
                FROM tasks
                WHERE user_id = ? AND status NOT IN ('удалено', 'исполнено')
                ORDER BY id DESC 
                LIMIT 20
            """, (executor,))
        
        tasks = cursor.fetchall()

        keyboard = InlineKeyboardMarkup(row_width=1)
        for task_id, task_text, status in tasks:
            keyboard.add(InlineKeyboardButton(
                f"{task_text[:30]}... (🔹: {task_id}, 🔄: {status})", 
                callback_data=f"status_task_{task_id}"
            ))
        
        keyboard.add(InlineKeyboardButton("✏️ Ввести ID вручную", callback_data="status_manual_id"))
        
        await bot.send_message(
            chat_id=message_obj.chat.id,
            text=f"Задачи исполнителя {'Без исполнителя' if executor is None or str(executor).lower() == 'none' else executor}:",
            reply_markup=keyboard
        )
        
    except Exception as e:
        logger.error(f"Ошибка при получении задач: {e}")
        await bot.send_message(chat_id=message_obj.chat.id, text="⚠ Ошибка при получении задач")

@dp.callback_query_handler(lambda c: c.data.startswith("status_task_"), state=StatusUpdate.waiting_for_task_selection)
async def process_selected_task_status(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка выбранной задачи для изменения статуса"""
    task_id = callback_query.data.split("_")[2]  # Формат: status_task_123
    await state.update_data(task_id=task_id)
    await show_status_options(callback_query.message, task_id)  # Передаем task_id
    await StatusUpdate.waiting_for_status_choice.set()

@dp.callback_query_handler(lambda c: c.data == "status_manual_id", state=[StatusUpdate.waiting_for_executor, StatusUpdate.waiting_for_task_selection])
async def ask_for_manual_id_status(callback_query: types.CallbackQuery):
    """Пропускаем выбор исполнителя при ручном вводе"""
    await bot.send_message(chat_id=callback_query.from_user.id, text="✏️ Введите ID задачи:")
    await StatusUpdate.waiting_for_task_selection.set()

@dp.message_handler(state=StatusUpdate.waiting_for_task_selection)
async def process_manual_task_id_status(message: types.Message, state: FSMContext):
    """Обработка ручного ввода ID задачи для изменения статуса"""
    try:
        task_id = int(message.text)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM tasks WHERE id=?", (task_id,))
        if not cursor.fetchone():
            await bot.send_message(chat_id=message.from_user.id, text="⚠ Задача не найдена!")
            await state.finish()
            return
        
        await state.update_data(task_id=task_id)
        await show_status_options(message, task_id)  # Передаем task_id
        await StatusUpdate.waiting_for_status_choice.set()
    except ValueError:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Введите числовой ID задачи!")
        await state.finish()

async def show_status_options(message_obj, task_id):
    """Показать варианты статусов"""
    keyboard = InlineKeyboardMarkup(row_width=3)

    cursor = conn.cursor()
    cursor.execute("SELECT creator_id FROM tasks WHERE id=?", (task_id,))
    task_creator = cursor.fetchone()
    if int(task_creator[0]) == message_obj.chat.id or message_obj.chat.id in MODERATOR_USERS:
        statuses = ["новая", "в работе", "ожидает доклада", "исполнено", "удалено"]
    else:
        statuses = ["новая", "в работе", "ожидает доклада", "исполнено"]
    
    buttons = [InlineKeyboardButton(
        status, 
        callback_data=f"set_status_{task_id}_{status}"
    ) for status in statuses]
    keyboard.add(*buttons)
    await bot.send_message(chat_id=message_obj.chat.id, text="📌 Выберите новый статус:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("set_status_"), state=StatusUpdate.waiting_for_status_choice)
async def process_status_update(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка изменения статуса"""
    try:
        # Извлекаем task_id и новый статус из callback_data
        _, _, task_id, new_status = callback_query.data.split("_")
        
        cursor = conn.cursor()
      
        cursor.execute("SELECT creator_id, task_text FROM tasks WHERE id=?", (task_id,))
        result = cursor.fetchone()
        
        if result:
            creator, task_text = result
      
        cursor.execute("""
            INSERT INTO tasks_log (id, user_id, chat_id, task_text, status, deadline, creator_id)
            SELECT id, user_id, chat_id, task_text, status, deadline, creator_id 
            FROM tasks 
            WHERE id=?
        """, (task_id,))
        cursor.execute("UPDATE tasks SET status=?, chat_id=? WHERE id=?", (new_status, callback_query.from_user.id, task_id))
        conn.commit()
        
        await bot.send_message(chat_id=callback_query.from_user.id, text=f"✅ Статус задачи {task_id} изменен на '{new_status}'")
        
        if creator is not None and creator != str(callback_query.from_user.id) and new_status in ('исполнено', 'удалено'):
          await bot.send_message(
              chat_id=creator,
              text=f"✅ Статус задачи {task_id} ({task_text}) изменен на '{new_status}'"
          )

        await state.finish()
    except Exception as e:
        logger.error(f"Ошибка при изменении статуса: {e}")
        await bot.send_message(chat_id=callback_query.from_user.id, text="⚠ Ошибка при изменении статуса")
        await state.finish()

# ======================
# ИЗМЕНИТЬ ТЕКСТ ЗАДАЧИ
# ======================

class TaskTextEditing(StatesGroup):
    waiting_for_executor_filter = State()     # Выбор исполнителя для фильтрации задач
    waiting_for_task_selection = State()      # Выбор задачи из списка (отфильтрованных по исполнителю)
    waiting_for_task_id = State()             # Ввод ID задачи вручную
    waiting_for_choice = State()              # Выбор между полной заменой и дополнением текста
    waiting_for_replacement = State()         # Ввод нового текста (полная замена)
    waiting_for_append = State()              # Ввод текста для дополнения

@dp.message_handler(lambda message: message.text == "✏️ Изменить задачу")
async def text_edit_start(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return
    if message.chat.type != "private":
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
        return

    cursor = conn.cursor()
    # Если пользователь — модератор, показываем всех исполнителей, иначе – только исполнителей задач, созданных им
    if message.from_user.id in MODERATOR_USERS:
        cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE status NOT IN ('удалено','исполнено') LIMIT 20")
    else:
        cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE creator_id=? AND status NOT IN ('удалено','исполнено') LIMIT 20", (message.from_user.id,))
    executors = cursor.fetchall()

    if not executors:
        await bot.send_message(chat_id=message.from_user.id, text="❌ Нет задач для изменения")
        return

    keyboard = InlineKeyboardMarkup(row_width=2)
    buttons = []
    # Формируем кнопки для каждого исполнителя
    for (executor,) in executors:
        if executor:
            label = f"👤 {executor}"
            data = executor
        else:
            label = "👤 Без исполнителя"
            data = "none"
        buttons.append(InlineKeyboardButton(label, callback_data=f"text_edit_executor|{data}"))
    # Добавляем кнопки исполнителей одним вызовом, чтобы они распределились по рядам согласно row_width
    keyboard.add(*buttons)
    # Добавляем отдельный ряд для ручного ввода ID задачи
    keyboard.add(InlineKeyboardButton("✏️ Ввести ID задачи вручную", callback_data="text_edit_manual_id"))
    
    await bot.send_message(
        chat_id=message.from_user.id,
        text="Выберите исполнителя для фильтрации задач:",
        reply_markup=keyboard
    )
    await TaskTextEditing.waiting_for_executor_filter.set()

# Обработка выбора исполнителя из списка
@dp.callback_query_handler(lambda c: c.data.startswith("text_edit_executor|"), state=TaskTextEditing.waiting_for_executor_filter)
async def process_text_edit_executor(callback_query: types.CallbackQuery, state: FSMContext):
    # Извлекаем выбранного исполнителя
    executor = callback_query.data.split("|")[1]
    await state.update_data(executor=executor)
    
    # После выбора исполнителя выводим список задач, отфильтрованных по выбранному исполнителю
    cursor = conn.cursor()
    if executor.lower() == "none":
        if callback_query.from_user.id in MODERATOR_USERS:
            cursor.execute("SELECT id, task_text FROM tasks WHERE user_id IS NULL AND status NOT IN ('удалено','исполнено') LIMIT 20")
        else:
            cursor.execute("SELECT id, task_text FROM tasks WHERE user_id IS NULL AND creator_id=? AND status NOT IN ('удалено','исполнено') LIMIT 20", (callback_query.from_user.id,))
    else:
        if callback_query.from_user.id in MODERATOR_USERS:
            cursor.execute("SELECT id, task_text FROM tasks WHERE user_id=? AND status NOT IN ('удалено','исполнено') LIMIT 20", (executor,))
        else:
            cursor.execute("SELECT id, task_text FROM tasks WHERE user_id=? AND creator_id=? AND status NOT IN ('удалено','исполнено') LIMIT 20", (executor, callback_query.from_user.id))
    tasks = cursor.fetchall()
    
    if not tasks:
        await bot.send_message(chat_id=callback_query.from_user.id, text="❌ Нет задач для выбранного исполнителя.")
        await state.finish()
        return

    keyboard = InlineKeyboardMarkup(row_width=1)
    for task_id, task_text in tasks:
        preview = (task_text[:30] + "...") if len(task_text) > 30 else task_text
        keyboard.add(InlineKeyboardButton(f"🔹 {preview} (ID: {task_id})", callback_data=f"text_edit_task_{task_id}"))
    # Добавляем кнопку для ручного ввода ID задачи
    keyboard.add(InlineKeyboardButton("✏️ Ввести ID задачи вручную", callback_data="text_edit_manual_id"))
    
    await bot.send_message(
        chat_id=callback_query.message.chat.id,
        text="Выберите задачу для изменения текста:",
        reply_markup=keyboard
    )
    await TaskTextEditing.waiting_for_task_selection.set()
    await bot.answer_callback_query(callback_query.id)

@dp.callback_query_handler(lambda c: c.data.startswith("text_edit_task_"), state=TaskTextEditing.waiting_for_task_selection)
async def process_text_edit_task(callback_query: types.CallbackQuery, state: FSMContext):
    try:
        task_id = int(callback_query.data.split("_")[-1])
    except (ValueError, IndexError):
        await bot.send_message(chat_id=callback_query.from_user.id, text="⚠ Неверные данные задачи!")
        await state.finish()
        return

    cursor = conn.cursor()
    cursor.execute("SELECT task_text, creator_id FROM tasks WHERE id=?", (task_id,))
    result = cursor.fetchone()
    if not result:
        await bot.send_message(chat_id=callback_query.from_user.id, text="⚠ Задача не найдена!")
        await state.finish()
        return
    current_text, creator_id = result
    await state.update_data(task_id=task_id, old_text=current_text, creator_id=creator_id)
    
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("Полностью заменить", callback_data="text_edit_full"),
        InlineKeyboardButton("Дополнить текст", callback_data="text_edit_append")
    )
    await bot.send_message(
        chat_id=callback_query.message.chat.id,
        text=f"<b>Текущий текст задачи:</b>\n{current_text}\n\nВыберите действие:",
        reply_markup=keyboard,
        parse_mode=ParseMode.HTML
    )
    await TaskTextEditing.waiting_for_choice.set()
    await bot.answer_callback_query(callback_query.id)

# Обработка ввода ID задачи вручную (на шаге выбора задачи)
@dp.callback_query_handler(lambda c: c.data == "text_edit_manual_id", state=[TaskTextEditing.waiting_for_executor_filter, TaskTextEditing.waiting_for_task_selection])
async def ask_manual_text_id(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.send_message(chat_id=callback_query.from_user.id, text="✏️ Введите ID задачи:")
    await TaskTextEditing.waiting_for_task_id.set()
    await bot.answer_callback_query(callback_query.id)

@dp.message_handler(state=TaskTextEditing.waiting_for_task_id)
async def process_task_id_text_edit(message: types.Message, state: FSMContext):
    try:
        task_id = int(message.text.strip())
    except ValueError:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Введите корректный числовой ID задачи!")
        await state.finish()
        return

    cursor = conn.cursor()
    cursor.execute("SELECT task_text, creator_id FROM tasks WHERE id=?", (task_id,))
    result = cursor.fetchone()
    if not result:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Задача не найдена!")
        await state.finish()
        return
    current_text, creator_id = result
    await state.update_data(task_id=task_id, old_text=current_text, creator_id=creator_id)
    
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("Полностью заменить", callback_data="text_edit_full"),
        InlineKeyboardButton("Дополнить текст", callback_data="text_edit_append")
    )
    await bot.send_message(chat_id=message.from_user.id,
                           text=f"Текущий текст задачи:\n{current_text}\n\nВыберите действие:",
                           reply_markup=keyboard)
    await TaskTextEditing.waiting_for_choice.set()

@dp.callback_query_handler(lambda c: c.data == "text_edit_full", state=TaskTextEditing.waiting_for_choice)
async def process_text_edit_choice_full(callback_query: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    creator_id = data.get("creator_id")
    # Полная замена разрешена только если пользователь – создатель задачи или модератор
    if int(creator_id) != callback_query.from_user.id and callback_query.from_user.id not in MODERATOR_USERS:
        await bot.send_message(callback_query.from_user.id,
                               text="⚠ Полная замена текста доступна только создателю задачи или модераторам!")
        await state.finish()
        return
    await bot.send_message(callback_query.from_user.id,
                           text="Введите новый текст задачи (старый текст будет полностью заменен):")
    await TaskTextEditing.waiting_for_replacement.set()
    await bot.answer_callback_query(callback_query.id)

@dp.message_handler(state=TaskTextEditing.waiting_for_replacement)
async def process_text_replacement(message: types.Message, state: FSMContext):
    new_text = message.text.strip()
    data = await state.get_data()
    task_id = data.get("task_id")
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO tasks_log (id, user_id, chat_id, task_text, status, deadline, creator_id)
            SELECT id, user_id, chat_id, task_text, status, deadline, creator_id
            FROM tasks WHERE id=?
        """, (task_id,))
        cursor.execute("UPDATE tasks SET task_text=?, chat_id=? WHERE id=?", (new_text, message.from_user.id, task_id))
        conn.commit()
        await bot.send_message(message.chat.id, text=f"✅ Текст задачи {task_id} успешно обновлен.")
    except Exception as e:
        logger.error(f"Ошибка при обновлении текста задачи: {e}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Ошибка при обновлении текста задачи.")
    finally:
        await state.finish()

@dp.callback_query_handler(lambda c: c.data == "text_edit_append", state=TaskTextEditing.waiting_for_choice)
async def process_text_edit_choice_append(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.send_message(callback_query.from_user.id,
                           text="Введите текст, который необходимо добавить в конец текущего описания:")
    await TaskTextEditing.waiting_for_append.set()
    await bot.answer_callback_query(callback_query.id)

@dp.message_handler(state=TaskTextEditing.waiting_for_append)
async def process_text_append(message: types.Message, state: FSMContext):
    append_text = message.text.strip()
    data = await state.get_data()
    task_id = data.get("task_id")
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT task_text FROM tasks WHERE id=?", (task_id,))
        result = cursor.fetchone()
        if not result:
            await bot.send_message(chat_id=message.from_user.id, text="⚠ Задача не найдена!")
            await state.finish()
            return
        old_text = result[0]
        new_text = old_text + "\n" + append_text
        cursor.execute("""
            INSERT INTO tasks_log (id, user_id, chat_id, task_text, status, deadline, creator_id)
            SELECT id, user_id, chat_id, task_text, status, deadline, creator_id
            FROM tasks WHERE id=?
        """, (task_id,))
        cursor.execute("UPDATE tasks SET task_text=?, chat_id=? WHERE id=?", (new_text, message.from_user.id, task_id))
        conn.commit()
        await bot.send_message(chat_id=message.from_user.id, text=f"✅ Текст задачи {task_id} успешно дополнен.")
    except Exception as e:
        logger.error(f"Ошибка при дополнении текста задачи: {e}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Ошибка при дополнении текста задачи.")
    finally:
        await state.finish()

# ======================
# ИЗМЕНЕНИЕ ИСПОЛНИТЕЛЯ
# ======================

class ExecutorUpdate(StatesGroup):
    waiting_for_executor = State()
    waiting_for_task_selection = State()
    waiting_for_new_executor = State()

@dp.message_handler(lambda message: message.text == "👤 Изменить исполнителя")
async def executor_select_task(message: types.Message):
    """Начало процесса изменения исполнителя"""
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return

    if message.chat.type != "private":
      await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
      return
    
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE status NOT IN ('удалено', 'исполнено') LIMIT 20")
    executors = cursor.fetchall()
    
    if not executors:
        await message.reply("❌ Нет задач для изменения исполнителя")
        return

    # Создаем inline-клавиатуру с исполнителями
    keyboard = InlineKeyboardMarkup(row_width=2)
    for i in range(0, len(executors), 2):
        row = executors[i:i+2]
        row_buttons = [
            InlineKeyboardButton(
                f"👤 {executor[0] if executor[0] else 'Без исполнителя'}",
                callback_data=f"executor_filter|{executor[0]}"
            ) for executor in row
        ]
        keyboard.add(*row_buttons)
    
    keyboard.add(InlineKeyboardButton("✏️ Ввести ID задачи", callback_data="executor_manual_id"))
    await message.reply("Выберите исполнителя для фильтрации задач:", reply_markup=keyboard)
    await ExecutorUpdate.waiting_for_executor.set()

@dp.callback_query_handler(lambda c: c.data.startswith("executor_filter|"), state=ExecutorUpdate.waiting_for_executor)
async def process_executor_filter(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка выбора исполнителя для фильтрации"""
    executor = callback_query.data.split("|")[-1]
    await state.update_data(executor=executor)
    await show_executor_tasks(callback_query.message, executor)
    await ExecutorUpdate.waiting_for_task_selection.set()

async def show_executor_tasks(message_obj, executor):
    """Отображение задач выбранного исполнителя"""
    try:
        cursor = conn.cursor()

        if message_obj.chat.id in MODERATOR_USERS:
            if executor.lower() == "none":
                cursor.execute("""
                    SELECT id, task_text, status 
                    FROM tasks
                    WHERE user_id IS NULL AND status NOT IN ('удалено', 'исполнено')
                    ORDER BY id DESC 
                    LIMIT 20
                """)
            else:
                cursor.execute("""
                    SELECT id, task_text, status 
                    FROM tasks
                    WHERE user_id = ? AND status NOT IN ('удалено', 'исполнено')
                    ORDER BY id DESC 
                    LIMIT 20
                """, (executor,))
        else:
            if executor.lower() == "none":
                cursor.execute("""
                    SELECT id, task_text, status 
                    FROM tasks
                    WHERE user_id IS NULL AND status NOT IN ('удалено', 'исполнено') AND creator_id=?
                    ORDER BY id DESC 
                    LIMIT 20
                """, (str(message_obj.chat.id),))
            else:
                cursor.execute("""
                    SELECT id, task_text, status 
                    FROM tasks
                    WHERE user_id = ? AND status NOT IN ('удалено', 'исполнено') AND creator_id=?
                    ORDER BY id DESC 
                    LIMIT 20
                """, (executor, str(message_obj.chat.id)))

        tasks = cursor.fetchall()

        keyboard = InlineKeyboardMarkup(row_width=1)
        for task_id, task_text, current_executor in tasks:
            keyboard.add(InlineKeyboardButton(
                f"{task_text[:30]}... (ID: {task_id})", 
                callback_data=f"executor_task_{task_id}"
            ))

        keyboard.add(InlineKeyboardButton("✏️ Ввести ID вручную", callback_data="executor_manual_id"))
        await bot.send_message(
            chat_id=message_obj.chat.id,
            text=f"Задачи исполнителя {'Без исполнителя' if executor is None or str(executor).lower() == 'none' else executor}:",
            reply_markup=keyboard
        )
    except Exception as e:
        logger.error(f"Ошибка при получении задач: {e}")

@dp.callback_query_handler(lambda c: c.data.startswith("executor_task_"), state=ExecutorUpdate.waiting_for_task_selection)
async def process_selected_task_executor(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка выбранной задачи"""
    task_id = callback_query.data.split("_")[2]
    await state.update_data(task_id=task_id)
    
    # Получаем список исполнителей для inline-клавиатуры
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE status<>'удалено' LIMIT 20")
    executors = cursor.fetchall()
    
    # Создаем inline-клавиатуру
    keyboard = InlineKeyboardMarkup(row_width=2)
    buttons = []
    for executor in executors:
        if executor[0]:
            buttons.append(InlineKeyboardButton(
                executor[0], 
                callback_data=f"executor_choice|{executor[0]}"
            ))
    
    # Добавляем кнопку ручного ввода
    keyboard.add(*buttons)
    keyboard.row(InlineKeyboardButton(
        "✏️ Ввести вручную", 
        callback_data="executor_manual_input"
    ))
    
    await bot.send_message(
        chat_id=callback_query.from_user.id,
        text="👤 Выберите нового исполнителя:",
        reply_markup=keyboard
    )
    await ExecutorUpdate.waiting_for_new_executor.set()

@dp.callback_query_handler(
    lambda c: c.data == "executor_manual_id", 
    state=[ExecutorUpdate.waiting_for_executor, ExecutorUpdate.waiting_for_task_selection]
)
async def ask_for_manual_id_executor(callback_query: types.CallbackQuery):
    """Обработка ручного ввода ID задачи"""
    await bot.send_message(chat_id=callback_query.from_user.id, text="✏️ Введите ID задачи:")
    await ExecutorUpdate.waiting_for_task_selection.set()

@dp.message_handler(state=ExecutorUpdate.waiting_for_task_selection)
async def process_manual_task_id_executor(message: types.Message, state: FSMContext):
    """Обработка ручного ввода ID задачи"""
    try:
        task_id = int(message.text)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM tasks WHERE id=?", (task_id,))
        if not cursor.fetchone():
            await bot.send_message(chat_id=message.from_user.id, text="⚠ Задача не найдена!")
            await state.finish()
            return
        
        await state.update_data(task_id=task_id)
        
        # Повторно используем логику создания inline-клавиатуры
        cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE status<>'удалено' LIMIT 20")
        executors = cursor.fetchall()
        
        keyboard = InlineKeyboardMarkup(row_width=2)
        buttons = []
        for executor in executors:
            if executor[0]:
                buttons.append(InlineKeyboardButton(
                    executor[0], 
                    callback_data=f"executor_choice|{executor[0]}"
                ))
        
        keyboard.add(*buttons)
        keyboard.row(InlineKeyboardButton(
            "✏️ Ввести вручную", 
            callback_data="executor_manual_input"
        ))
        
        await bot.send_message(
            chat_id=message.from_user.id,
            text="👤 Выберите нового исполнителя:",
            reply_markup=keyboard
        )
        await ExecutorUpdate.waiting_for_new_executor.set()
        
    except ValueError:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Введите числовой ID задачи!")
        await state.finish()

@dp.callback_query_handler(
    lambda c: c.data.startswith("executor_choice|"), 
    state=ExecutorUpdate.waiting_for_new_executor
)
async def process_executor_choice(callback: types.CallbackQuery, state: FSMContext):
    """Обработка выбора исполнителя из списка"""
    new_executor = callback.data.split("|")[-1]
    await process_and_save_executor(callback.message, new_executor, state)

@dp.callback_query_handler(
    lambda c: c.data == "executor_manual_input", 
    state=ExecutorUpdate.waiting_for_new_executor
)
async def ask_manual_executor_input(callback: types.CallbackQuery):
    """Запрос ручного ввода исполнителя"""
    await bot.send_message(callback.from_user.id, "✏️ Введите @username")
    await ExecutorUpdate.waiting_for_new_executor.set()

@dp.message_handler(state=ExecutorUpdate.waiting_for_new_executor)
async def process_new_executor(message: types.Message, state: FSMContext):
    """Обработка ручного ввода исполнителя"""
    await process_and_save_executor(message, message.text.strip(), state)

async def process_and_save_executor(message_obj, new_executor: str, state: FSMContext):
    """Общая логика сохранения нового исполнителя"""
    try:
        user_data = await state.get_data()
        task_id = user_data['task_id']
        chat_type = message_obj.chat.type
      
        cursor = conn.cursor()

        cursor.execute("SELECT creator_id FROM tasks WHERE id=?", (task_id,))
        task_creator = cursor.fetchone()
        if int(task_creator[0]) != message_obj.chat.id and message_obj.chat.id not in MODERATOR_USERS:
            await bot.send_message(chat_id=message_obj.chat.id, text="⚠ Вы не можете изменить эту задачу!")
            await state.finish()
            return
          
        cursor.execute("""
            INSERT INTO tasks_log (id, user_id, chat_id, task_text, status, deadline, creator_id)
            SELECT id, user_id, chat_id, task_text, status, deadline, creator_id
            FROM tasks 
            WHERE id=?
        """, (task_id,))
        cursor.execute("UPDATE tasks SET user_id=?, chat_id=? WHERE id=?", (new_executor, message_obj.chat.id, task_id))
        conn.commit()

        reply_markup = menu_keyboard if chat_type == "private" else group_menu_keyboard
        await bot.send_message(
            chat_id=message_obj.chat.id,
            text=f"✅ Исполнитель задачи {task_id} изменен на '{new_executor}'",
            reply_markup=reply_markup
        )
        await state.finish()
        
    except Exception as e:
        logger.error(f"Ошибка при изменении исполнителя: {e}")
        await bot.send_message(chat_id=message_obj.chat.id, text="⚠ Ошибка при изменении исполнителя")
        await state.finish()
      
# ======================
# ИЗМЕНЕНИЕ СРОКА
# ======================

class TaskUpdate(StatesGroup):
    waiting_for_executor = State()
    waiting_for_task_selection = State()
    waiting_for_deadline_choice = State()
    waiting_for_custom_deadline = State()

@dp.message_handler(lambda message: message.text == "⏳ Изменить срок")
async def deadline_select_task(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return

    if message.chat.type != "private":
      await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
      return
    
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_id FROM tasks WHERE status NOT IN ('удалено', 'исполнено') LIMIT 20")
    executors = cursor.fetchall()
    
    if not executors:
        await message.reply("❌ Нет задач для изменения срока")
        return

    keyboard = InlineKeyboardMarkup(row_width=2)
    for i in range(0, len(executors), 2):
        row = executors[i:i+2]
        row_buttons = [
            InlineKeyboardButton(
                f"👤 {executor[0] if executor[0] else 'Без исполнителя'}",
                callback_data=f"deadline_filter|{executor[0]}"
            ) for executor in row
        ]
        keyboard.add(*row_buttons)
    
    keyboard.add(InlineKeyboardButton("✏️ Ввести ID задачи", callback_data="deadline_manual_id"))
    await message.reply("Выберите исполнителя для фильтрации задач:", reply_markup=keyboard)
    await TaskUpdate.waiting_for_executor.set()

@dp.callback_query_handler(lambda c: c.data.startswith("deadline_filter|"), state=TaskUpdate.waiting_for_executor)
async def process_deadline_filter(callback_query: types.CallbackQuery, state: FSMContext):
    executor = callback_query.data.split("|")[-1]
    await state.update_data(executor=executor)
    await show_deadline_tasks(callback_query.message, executor)
    await TaskUpdate.waiting_for_task_selection.set()

async def show_deadline_tasks(message_obj, executor):
    try:
        cursor = conn.cursor()
        if message_obj.chat.id in MODERATOR_USERS:
            if executor.lower() == "none":
                cursor.execute("""
                    SELECT id, task_text, deadline 
                    FROM tasks
                    WHERE user_id IS NULL AND status NOT IN ('удалено', 'исполнено')
                    ORDER BY id DESC 
                    LIMIT 20
                """)
            else:
                cursor.execute("""
                    SELECT id, task_text, deadline 
                    FROM tasks
                    WHERE user_id = ? AND status NOT IN ('удалено', 'исполнено')
                    ORDER BY id DESC 
                    LIMIT 20
                """, (executor,))
        else:
            if executor.lower() == "none":
                cursor.execute("""
                    SELECT id, task_text, deadline 
                    FROM tasks
                    WHERE user_id IS NULL AND status NOT IN ('удалено', 'исполнено') AND creator_id=?
                    ORDER BY id DESC 
                    LIMIT 20
                """, (str(message_obj.chat.id),))
            else:
                cursor.execute("""
                    SELECT id, task_text, deadline 
                    FROM tasks
                    WHERE user_id = ? AND status NOT IN ('удалено', 'исполнено') AND creator_id=?
                    ORDER BY id DESC 
                    LIMIT 20
                """, (executor, str(message_obj.chat.id)))
        
        tasks = cursor.fetchall()

        keyboard = InlineKeyboardMarkup(row_width=1)
        for task_id, task_text, deadline in tasks:
            keyboard.add(InlineKeyboardButton(
                f"{task_text[:30]}... (ID: {task_id})", 
                callback_data=f"deadline_task_{task_id}"
            ))

        keyboard.add(InlineKeyboardButton("✏️ Ввести ID вручную", callback_data="deadline_manual_id"))
        await bot.send_message(
            chat_id=message_obj.chat.id,
            text=f"Задачи исполнителя {'Без исполнителя' if executor is None or str(executor).lower() == 'none' else executor}:",
            reply_markup=keyboard
        )
    except Exception as e:
        logger.error(f"Ошибка при получении задач: {e}")

@dp.callback_query_handler(lambda c: c.data.startswith("deadline_task_"), state=TaskUpdate.waiting_for_task_selection)
async def process_selected_task(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка выбранной задачи"""
    task_id = callback_query.data.split("_")[-1]
    await state.update_data(task_id=task_id)
    await show_deadline_options(callback_query.message)
    await TaskUpdate.waiting_for_deadline_choice.set()

@dp.callback_query_handler(
    lambda c: c.data == "deadline_manual_id", 
    state=[TaskUpdate.waiting_for_executor, TaskUpdate.waiting_for_task_selection]  # Добавить оба состояния
)
async def ask_for_manual_id(callback_query: types.CallbackQuery):
    await bot.send_message(chat_id=callback_query.from_user.id, text="✏️ Введите ID задачи:")
    await TaskUpdate.waiting_for_task_selection.set()

@dp.message_handler(state=TaskUpdate.waiting_for_task_selection)
async def process_manual_task_id(message: types.Message, state: FSMContext):
    """Обработка ручного ввода ID задачи"""
    try:
        task_id = int(message.text)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM tasks WHERE id=?", (task_id,))
        if not cursor.fetchone():
            await bot.send_message(chat_id=message.from_user.id, text="⚠ Задача не найдена!")
            return
        
        await state.update_data(task_id=task_id)
        await show_deadline_options(message)
        await TaskUpdate.waiting_for_deadline_choice.set()
    except ValueError:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Введите числовой ID задачи!")
        await state.finish()

async def show_deadline_options(message_obj):
    """Показать варианты выбора срока"""
    keyboard = get_deadline_keyboard(with_none_option=True)
    await bot.send_message(chat_id=message_obj.chat.id, text="⏳ Выберите новый срок:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("set_deadline_"), state=TaskUpdate.waiting_for_deadline_choice)
async def process_deadline_choice(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработка выбора типа срока"""
    if callback_query.data == "set_deadline_custom":
        await bot.send_message(chat_id=callback_query.from_user.id, text="📅 Введите дату в формате DD.MM.YYYY:")
        await TaskUpdate.waiting_for_custom_deadline.set()
    else:
        user_data = await state.get_data()
        task_id = user_data['task_id']
        
        if callback_query.data == "set_deadline_none":
            new_deadline = None
            response = "✅ Срок выполнения удален"
        else:
            new_deadline = callback_query.data.split("_")[2]
            response = f"✅ Новый срок: {new_deadline}"
        
        cursor = conn.cursor()

        cursor.execute("SELECT creator_id FROM tasks WHERE id=?", (task_id,))
        task_creator = cursor.fetchone()
        if int(task_creator[0]) != callback_query.from_user.id and callback_query.from_user.id not in MODERATOR_USERS:
            await bot.send_message(chat_id=callback_query.from_user.id, text="⚠ Вы не можете изменить эту задачу!")
            await state.finish()
            return
          
        cursor.execute("""
            INSERT INTO tasks_log (id, user_id, chat_id, task_text, status, deadline, creator_id)
            SELECT id, user_id, chat_id, task_text, status, deadline, creator_id
            FROM tasks 
            WHERE id=?
        """, (task_id,))
        cursor.execute("UPDATE tasks SET deadline=?, chat_id=? WHERE id=?", (new_deadline, callback_query.from_user.id, task_id))
        conn.commit()
        
        await bot.send_message(chat_id=callback_query.from_user.id, text=response)
        await state.finish()

@dp.message_handler(state=TaskUpdate.waiting_for_custom_deadline)
async def process_custom_deadline(message: types.Message, state: FSMContext):
    """Обработка ввода даты вручную"""
    try:
        # Попытка парсинга по первому шаблону
        try:
            dt = datetime.strptime(message.text.strip(), "%Y-%m-%d %H:%M")
        except ValueError:
            try:
                dt = datetime.strptime(message.text.strip(), "%d.%m.%Y %H:%M")
            except ValueError:
                try:
                    dt = datetime.strptime(message.text.strip(), "%d.%m.%y %H:%M")
                except ValueError:
                    try:
                        dt = datetime.strptime(message.text.strip(), "%Y-%m-%d")
                    except ValueError:
                        # Если не получилось, пробуем другой формат
                        try:
                            dt = datetime.strptime(message.text.strip(), "%d.%m.%Y")
                        except ValueError:
                            dt = datetime.strptime(message.text.strip(), "%d.%m.%y")
        new_deadline = dt.strftime("%Y-%m-%d %H:%M")
        
        user_data = await state.get_data()
        task_id = user_data['task_id']
        
        cursor = conn.cursor()

        cursor.execute("SELECT creator_id FROM tasks WHERE id=?", (task_id,))
        task_creator = cursor.fetchone()
        if int(task_creator[0]) != message.from_user.id and message.from_user.id not in MODERATOR_USERS:
            await bot.send_message(chat_id=message.from_user.id, text="⚠ Вы не можете изменить эту задачу!")
            await state.finish()
            return
  
        cursor.execute("""
            INSERT INTO tasks_log (id, user_id, chat_id, task_text, status, deadline, creator_id)
            SELECT id, user_id, chat_id, task_text, status, deadline, creator_id
            FROM tasks 
            WHERE id=?
        """, (task_id,))
        cursor.execute("UPDATE tasks SET deadline=?, chat_id=? WHERE id=?", (new_deadline, message.from_user.id, task_id))
        conn.commit()
        
        await bot.send_message(chat_id=message.from_user.id,text=f"✅ Новый срок установлен: {new_deadline}")
        await state.finish()
    except ValueError:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Неверный формат даты! Используйте DD.MM.YYYY")
        await state.finish()

# ======================
# СПИСОК ЗАДАЧ
# ======================

current_page = {}
current_filters = {}

@dp.message_handler(lambda message: message.text == "📋 Список задач")
async def list_tasks(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return

    if message.chat.type != "private":
      await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
      return

    """Просмотр списка задач с выбором исполнителя и пагинацией"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT user_id FROM tasks 
            WHERE status NOT IN ('удалено', 'исполнено')
            LIMIT 20
        """)
        executors = cursor.fetchall()
        if not executors:
            await message.reply("❌ Нет задач для отображения")
            return
        keyboard = InlineKeyboardMarkup(row_width=2)
        for i in range(0, len(executors), 2):
            row = executors[i:i+2]
            row_buttons = [
                InlineKeyboardButton(
                    f"👤 {executor[0] if executor[0] else 'Без исполнителя'}",
                    callback_data=f"listtasks_executor|{executor[0]}"
                ) for executor in row
            ]
            keyboard.add(*row_buttons)
        await message.reply("Выберите исполнителя для фильтрации задач:", reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Ошибка при получении списка задач: {str(e)}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Ошибка при получении списка задач.")

@dp.callback_query_handler(lambda c: c.data.startswith("listtasks_executor|"))
async def process_listtasks_executor(callback_query: types.CallbackQuery):
    executor = callback_query.data.split("|")[1]
    user_id = callback_query.from_user.id
    current_page[user_id] = 0
    current_filters[user_id] = executor  # Сохраняем фильтр
    sent_message = await show_tasks_page(callback_query.message, user_id, page=0, executor_filter=executor)
    current_page[f"{user_id}_message_id"] = sent_message.message_id
    await bot.answer_callback_query(callback_query.id)


async def show_tasks_page(message: types.Message, user_id: int, page: int, executor_filter: str = None):
    try:
        cursor = conn.cursor()
        # Если указан фильтр по исполнителю, добавляем условие
        if executor_filter and executor_filter.lower() == "none":
            cursor.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('удалено','исполнено') AND user_id IS NULL")
        elif executor_filter:
            cursor.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('удалено','исполнено') AND user_id = ?", (executor_filter,))
        else:
            cursor.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('удалено','исполнено')")
        total_tasks = cursor.fetchone()[0]
        
        if total_tasks == 0:
            return await bot.send_message(message.chat.id, "📭 Нет активных задач.")
        
        total_pages = (total_tasks - 1) // 10
        page = max(0, min(page, total_pages))
        
        # Получаем задачи с учетом фильтра
        if executor_filter:
            if executor_filter.lower() == "none":
                cursor.execute("""
                    SELECT id, user_id, task_text, status, deadline 
                    FROM tasks 
                    WHERE status NOT IN ('удалено','исполнено') AND user_id IS NULL
                    ORDER BY datetime(deadline) ASC, id ASC
                    LIMIT 10 OFFSET ?
                """, (page * 10,))
            else:
                cursor.execute("""
                    SELECT id, user_id, task_text, status, deadline 
                    FROM tasks 
                    WHERE status NOT IN ('удалено','исполнено') AND user_id = ?
                    ORDER BY datetime(deadline) ASC, id ASC
                    LIMIT 10 OFFSET ?
                """, (executor_filter, page * 10))
        else:
            cursor.execute("""
                SELECT id, user_id, task_text, status, deadline 
                FROM tasks 
                WHERE status NOT IN ('удалено','исполнено')
                ORDER BY datetime(deadline) ASC, id ASC
                LIMIT 10 OFFSET ?
            """, (page * 10,))
        tasks = cursor.fetchall()

        result = []
        for task in tasks:
            task_id, task_user, task_text, status, deadline = task
            result.append(
                f"🔹: {task_id} 📝: {task_text}\n\n"
                f"🔄: {status} ⏳: {format_date(deadline) if deadline else 'нет срока'}\n"
                f"──────────"
            )
        keyboard = InlineKeyboardMarkup(row_width=3)
        buttons = []
        if page > 0:
            buttons.append(InlineKeyboardButton("⬅️ Назад", callback_data=f"tasks_prev_{page-1}"))
        buttons.append(InlineKeyboardButton(f"{page+1}/{total_pages+1}", callback_data="tasks_page"))
        if page < total_pages:
            buttons.append(InlineKeyboardButton("Вперед ➡️", callback_data=f"tasks_next_{page+1}"))
        keyboard.row(*buttons)
        
        header = f"📋 Список задач (страница {page+1} из {total_pages+1})"
        if executor_filter:
            executor_display = 'Без исполнителя' if str(executor_filter).lower() == 'none' else executor_filter
            header = f"📋 Задачи для 👤: <b>{executor_display}</b> (страница {page+1} из {total_pages+1})"
        sent_message = await bot.send_message(
            chat_id=message.chat.id,
            text=header + ":\n\n" + "\n".join(result),
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        return sent_message
        
    except Exception as e:
        logger.error(f"Ошибка при отображении страницы задач: {str(e)}")
        await bot.send_message(message.from_user.id, "⚠ Ошибка при отображении задач.")
        return None

@dp.callback_query_handler(lambda c: c.data.startswith(("tasks_prev_", "tasks_next_")))
async def process_tasks_pagination(callback_query: types.CallbackQuery):
    """Обработка переключения страниц"""
    try:
        user_id = callback_query.from_user.id
        action, page = callback_query.data.split("_")[1:3]
        page = int(page)
        
        # Получаем сохраненный фильтр
        executor_filter = current_filters.get(user_id)
        
        current_page[user_id] = page
        
        class FakeMessage:
            def __init__(self, chat_id):
                self.chat = type('Chat', (), {'id': chat_id})()
                self.from_user = type('User', (), {'id': user_id})()
        
        fake_message = FakeMessage(callback_query.message.chat.id)
        
        # Передаем сохраненный фильтр
        sent_message = await show_tasks_page(fake_message, user_id, page, executor_filter)

        try:
            prev_message_id = current_page.get(f"{user_id}_message_id")
            if prev_message_id:
                await bot.delete_message(chat_id=callback_query.message.chat.id, message_id=prev_message_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение: {e}")
        
        if sent_message:
            current_page[f"{user_id}_message_id"] = sent_message.message_id
        
        await bot.answer_callback_query(callback_query.id)
        
    except Exception as e:
        logger.error(f"Ошибка при переключении страниц: {str(e)}")
        await bot.answer_callback_query(callback_query.id, "⚠ Ошибка при переключении страниц", show_alert=False)

# ======================
# СПИСОК ЗАДАЧ (по сроку)
# ======================

current_page_deadline = {}
current_filters_deadline = {}

def format_deadline_time(deadline_str):
    try:
        # Сначала пробуем распарсить срок как дату со временем
        dt = datetime.strptime(deadline_str, "%Y-%m-%d %H:%M")
    except Exception:
        return ""

    # Если время задано (не 00:00), возвращаем только время
    if dt.hour != 0 or dt.minute != 0:
        return dt.strftime("%H:%M")
    else:
        return ""

@dp.message_handler(lambda message: message.text == "📋 Список (по сроку)")
async def list_tasks_by_deadline(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return

    if message.chat.type != "private":
      await bot.send_message(chat_id=message.from_user.id, text="⛔ Команда для ЛС!")
      return

    """Просмотр списка задач с выбором срока и пагинацией"""
    try:
        cursor = conn.cursor()
        # Получаем уникальные сроки. Если срок отсутствует (NULL), то можно отобразить вариант "Без срока"
        cursor.execute("""
            SELECT DISTINCT date(deadline) deadline FROM tasks 
            WHERE status NOT IN ('удалено', 'исполнено')
            ORDER BY datetime(deadline) ASC
            LIMIT 20
        """)
        deadlines = cursor.fetchall()
        if not deadlines:
            await message.reply("❌ Нет задач для отображения")
            return
        keyboard = InlineKeyboardMarkup(row_width=2)
        for i in range(0, len(deadlines), 2):
            row = deadlines[i:i+2]
            row_buttons = []
            for d in row:
                if d[0]:
                    btn_text = format_date(d[0])
                    btn_data = d[0]
                else:
                    btn_text = "Без срока"
                    btn_data = "none"
                row_buttons.append(InlineKeyboardButton(
                    f"⏳ {btn_text}",
                    callback_data=f"listtasks_deadline|{btn_data}"
                ))
            keyboard.add(*row_buttons)
        await message.reply("Выберите срок для фильтрации задач:", reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Ошибка при получении списка сроков: {str(e)}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Ошибка при получении списка задач.")

@dp.callback_query_handler(lambda c: c.data.startswith("listtasks_deadline|"))
async def process_listtasks_deadline(callback_query: types.CallbackQuery):
    deadline_filter = callback_query.data.split("|")[1]
    user_id = callback_query.from_user.id
    current_page_deadline[user_id] = 0
    current_filters_deadline[user_id] = deadline_filter  # Сохраняем выбранный срок
    sent_message = await show_tasks_page_by_deadline(callback_query.message, user_id, page=0, deadline_filter=deadline_filter)
    current_page_deadline[f"{user_id}_message_id"] = sent_message.message_id
    await bot.answer_callback_query(callback_query.id)

async def show_tasks_page_by_deadline(message: types.Message, user_id: int, page: int, deadline_filter: str = None):
    try:
        cursor = conn.cursor()
        # Если выбран конкретный срок, считаем задачи с этим сроком.
        # Если выбран вариант "Без срока" (deadline_filter == "none"), ищем записи с deadline IS NULL.
        if deadline_filter and deadline_filter.lower() == "none":
            cursor.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('удалено','исполнено') AND deadline IS NULL")
        elif deadline_filter:
            cursor.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('удалено','исполнено') AND date(deadline) = ?", (deadline_filter,))
        else:
            cursor.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('удалено','исполнено')")
        total_tasks = cursor.fetchone()[0]
        
        if total_tasks == 0:
            return await bot.send_message(message.chat.id, "📭 Нет активных задач.")
        
        total_pages = (total_tasks - 1) // 10
        page = max(0, min(page, total_pages))
        
        # Получаем задачи с применённой фильтрацией по сроку
        if deadline_filter:
            if deadline_filter.lower() == "none":
                cursor.execute("""
                    SELECT id, user_id, task_text, status, deadline 
                    FROM tasks 
                    WHERE status NOT IN ('удалено','исполнено') AND deadline IS NULL
                    ORDER BY datetime(deadline) ASC, id ASC
                    LIMIT 10 OFFSET ?
                """, (page * 10,))
            else:
                cursor.execute("""
                    SELECT id, user_id, task_text, status, deadline 
                    FROM tasks 
                    WHERE status NOT IN ('удалено','исполнено') AND date(deadline) = ?
                    ORDER BY datetime(deadline) ASC, id ASC 
                    LIMIT 10 OFFSET ?
                """, (deadline_filter, page * 10))
        else:
            cursor.execute("""
                SELECT id, user_id, task_text, status, deadline 
                FROM tasks 
                WHERE status NOT IN ('удалено','исполнено')
                ORDER BY datetime(deadline) ASC, id ASC
                LIMIT 10 OFFSET ?
            """, (page * 10,))
        tasks = cursor.fetchall()

        result = []
        for task in tasks:
            task_id, task_user, task_text, status, deadline = task
            result.append(
                f"🔹: {task_id} 📝: {task_text}\n\n"
                f"👤: {task_user} 🔄: {status} {'⏳: ' + format_deadline_time(deadline) if format_deadline_time(deadline).strip() else ''}\n"
                f"──────────"
            )
        keyboard = InlineKeyboardMarkup(row_width=3)
        buttons = []
        if page > 0:
            buttons.append(InlineKeyboardButton("⬅️ Назад", callback_data=f"tasks_prev_{page-1}"))
        buttons.append(InlineKeyboardButton(f"{page+1}/{total_pages+1}", callback_data="tasks_page"))
        if page < total_pages:
            buttons.append(InlineKeyboardButton("Вперед ➡️", callback_data=f"tasks_next_{page+1}"))
        keyboard.row(*buttons)
        
        header = f"📋 Список задач (страница {page+1} из {total_pages+1})"
        if deadline_filter:
            deadline_display = 'Без срока' if deadline_filter.lower() == 'none' else deadline_filter
            header = f"📋 Задачи со сроком: <b>⏳: {format_date(deadline_display)}</b> (страница {page+1} из {total_pages+1})"
        sent_message = await bot.send_message(
            chat_id=message.chat.id,
            text=header + ":\n\n" + "\n".join(result),
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        return sent_message
        
    except Exception as e:
        logger.error(f"Ошибка при отображении страницы задач: {str(e)}")
        await bot.send_message(message.from_user.id, "⚠ Ошибка при отображении задач.")
        return None

@dp.callback_query_handler(lambda c: c.data.startswith(("tasks_prev_", "tasks_next_")))
async def process_tasks_pagination_deadline(callback_query: types.CallbackQuery):
    """Обработка переключения страниц для фильтрации по сроку"""
    try:
        user_id = callback_query.from_user.id
        action, page = callback_query.data.split("_")[1:3]
        page = int(page)
        
        deadline_filter = current_filters_deadline.get(user_id)
        current_page_deadline[user_id] = page
        
        class FakeMessage:
            def __init__(self, chat_id):
                self.chat = type('Chat', (), {'id': chat_id})()
                self.from_user = type('User', (), {'id': user_id})()
        
        fake_message = FakeMessage(callback_query.message.chat.id)
        sent_message = await show_tasks_page_by_deadline(fake_message, user_id, page, deadline_filter)
        
        try:
            prev_message_id = current_page_deadline.get(f"{user_id}_message_id")
            if prev_message_id:
                await bot.delete_message(chat_id=callback_query.message.chat.id, message_id=prev_message_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение: {e}")
        
        if sent_message:
            current_page_deadline[f"{user_id}_message_id"] = sent_message.message_id
        
        await bot.answer_callback_query(callback_query.id)
        
    except Exception as e:
        logger.error(f"Ошибка при переключении страниц: {str(e)}")
        await bot.answer_callback_query(callback_query.id, "⚠ Ошибка при переключении страниц", show_alert=False)

# ======================
# ЭКСПОРТ ЗАДАЧ В CSV
# ======================

@dp.message_handler(lambda message: message.text == "📤 Экспорт задач")
async def export_tasks_to_csv(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    """Экспорт всех задач в CSV файл с кодировкой win1251"""
    try:
        cursor = conn.cursor()
        cursor.execute(""" SELECT t.id, 
                              CASE WHEN u.name IS NULL 
                                   THEN t.user_id 
                              ELSE u.name END "Исполнитель", 
                              t.task_text as "Задача", 
                              t.status as "Статус", 
                              t.deadline as "Срок"
                        FROM tasks t
                        LEFT JOIN users u ON t.user_id = u.username
                        WHERE status NOT IN ('удалено', 'исполнено')
                        ORDER BY user_id ASC, datetime(deadline) ASC, id ASC""")
        tasks = cursor.fetchall()
        
        if not tasks:
            await bot.send_message(chat_id=message.from_user.id, text="📭 В базе нет задач для экспорта.")
            return

        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "tasks_export"
        
        # Задаем заголовки
        headers = ['№', 'Исполнитель', 'Задача', 'Статус', 'Срок']
        ws.append(headers)
        
        # Определяем стили
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        header_font = Font(bold=True)
        
        # Применяем стили к заголовкам
        for col, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            # Центрирование для заголовка (по желанию)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Записываем данные
        for row_data in tasks:
            # Преобразуем значения в строки, если нужно
            row = [str(item) if item is not None else '' for item in row_data]
            ws.append(row)
        
        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 16

        # Преобразуем значение ячеек столбца "Срок" (столбец E) к datetime,
        # затем задаем нужный формат
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                if cell.value:
                    try:
                        # Сначала пытаемся преобразовать значение как дату со временем
                        date_value = datetime.strptime(str(cell.value), "%Y-%m-%d %H:%M")
                    except Exception:
                        try:
                            # Если не получилось, пробуем преобразовать как дату без времени
                            date_value = datetime.strptime(str(cell.value), "%Y-%m-%d")
                        except Exception:
                            # Если преобразование не удалось, оставляем значение без изменений
                            continue
                    cell.value = date_value
                    # Если время задано (не равно 00:00), устанавливаем формат с временем
                    if date_value.hour != 0 or date_value.minute != 0:
                        cell.number_format = 'DD.MM.YYYY HH:MM'
                    else:
                        cell.number_format = 'DD.MM.YYYY'

        # Применяем границы ко всем ячейкам
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border

        # Устанавливаем перенос слов для третьего столбца (используем Alignment)
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Сохраняем Excel в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Отправляем файл в Telegram (используем InputFile)
        from aiogram.types import InputFile
        excel_file = InputFile(output, filename="tasks_export.xlsx")
        await message.reply_document(document=excel_file)
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте задач в Excel: {str(e)}", exc_info=True)
        await bot.send_message(chat_id=message.from_user.id, text=f"⚠ Ошибка при создании файла экспорта: {str(e)}")

# ======================
# ЭКСПОРТ ЗАДАЧ В CSV (с исполненными)
# ======================

@dp.message_handler(lambda message: message.text == "📤 Экспорт (с исполненными)")
async def export_tasks_to_csv2(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Доступ запрещен")
        return  
    """Экспорт всех задач в CSV файл с кодировкой win1251"""
    try:
        cursor = conn.cursor()
        cursor.execute(""" SELECT t.id, 
                              CASE WHEN u.name IS NULL 
                                   THEN t.user_id 
                              ELSE u.name END "Исполнитель", 
                              t.task_text as "Задача", 
                              t.status as "Статус", 
                              t.deadline as "Срок"
                        FROM tasks t
                        LEFT JOIN users u ON t.user_id = u.username
                        WHERE status NOT IN ('удалено')
                        ORDER BY user_id ASC, datetime(deadline) ASC, id ASC""")
        tasks = cursor.fetchall()
        
        if not tasks:
            await bot.send_message(chat_id=message.from_user.id, text="📭 В базе нет задач для экспорта.")
            return

        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "tasks_export"
        
        # Задаем заголовки
        headers = ['№', 'Исполнитель', 'Задача', 'Статус', 'Срок']
        ws.append(headers)
        
        # Определяем стили
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        header_font = Font(bold=True)
        
        # Применяем стили к заголовкам
        for col, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            # Центрирование для заголовка (по желанию)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Записываем данные
        for row_data in tasks:
            # Преобразуем значения в строки, если нужно
            row = [str(item) if item is not None else '' for item in row_data]
            ws.append(row)
        
        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 16

        # Преобразуем значение ячеек столбца "Срок" (столбец E) к datetime,
        # затем задаем нужный формат
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                if cell.value:
                    try:
                        # Сначала пытаемся преобразовать значение как дату со временем
                        date_value = datetime.strptime(str(cell.value), "%Y-%m-%d %H:%M")
                    except Exception:
                        try:
                            # Если не получилось, пробуем преобразовать как дату без времени
                            date_value = datetime.strptime(str(cell.value), "%Y-%m-%d")
                        except Exception:
                            # Если преобразование не удалось, оставляем значение без изменений
                            continue
                    cell.value = date_value
                    # Если время задано (не равно 00:00), устанавливаем формат с временем
                    if date_value.hour != 0 or date_value.minute != 0:
                        cell.number_format = 'DD.MM.YYYY HH:MM'
                    else:
                        cell.number_format = 'DD.MM.YYYY'

        # Применяем границы ко всем ячейкам
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border

        # Устанавливаем перенос слов для третьего столбца (используем Alignment)
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Сохраняем Excel в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Отправляем файл в Telegram (используем InputFile)
        from aiogram.types import InputFile
        excel_file = InputFile(output, filename="tasks_export.xlsx")
        await message.reply_document(document=excel_file)
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте задач в Excel: {str(e)}", exc_info=True)
        await bot.send_message(chat_id=message.from_user.id, text=f"⚠ Ошибка при создании файла экспорта: {str(e)}")

# ======================
# ЭКСПОРТ ЗАДАЧ В CSV (с удаленными и историей изменений)
# ======================

@dp.message_handler(commands=["export3"])
async def export_tasks_to_csv3(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Только администратор может делать полный экспорт")
        return
      
    """Экспорт всех задач в CSV файл с кодировкой win1251"""
    try:
        cursor = conn.cursor()
        cursor.execute("""SELECT id, creator_id, user_id, chat_id, task_text, status, deadline, 999999 as "id_log" 
                          FROM tasks
                          UNION ALL SELECT id, creator_id, user_id, chat_id, task_text, status, deadline, id_log 
                          FROM tasks_log
                          ORDER BY id DESC, id_log DESC
                      """)
        tasks = cursor.fetchall()
        
        if not tasks:
            await bot.send_message(chat_id=message.from_user.id, text="📭 В базе нет задач для экспорта.")
            return

        # Создаем CSV в памяти
        output = io.BytesIO()
        
        # Используем TextIOWrapper с нужной кодировкой
        text_buffer = io.TextIOWrapper(
            output,
            encoding='utf-8-sig',
            errors='replace',  # заменяем некодируемые символы
            newline=''
        )
        
        writer = csv.writer(
            text_buffer,
            delimiter=';',  # Указываем нужный разделитель
            quoting=csv.QUOTE_MINIMAL
        )
        
        # Заголовки столбцов
        headers = ['ID', 'ID создателя', 'Исполнитель', 'ID редактора', 'Задача', 'Статус', 'Срок', 'ID Log']
        writer.writerow(headers)
        
        # Данные
        for task in tasks:
            # Преобразуем все значения в строки
            row = [
                str(item) if item is not None else ''
                for item in task
            ]
            writer.writerow(row)
        
        # Важно: закрыть TextIOWrapper перед использованием буфера
        text_buffer.flush()
        text_buffer.detach()  # Отсоединяем TextIOWrapper от BytesIO
        output.seek(0)
        
        # Создаем временный файл
        csv_file = InputFile(output, filename="tasks_export.csv")
        
        await message.reply_document(
            document=csv_file
        )
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте задач: {str(e)}", exc_info=True)
        await bot.send_message(chat_id=message.from_user.id,text=f"⚠ Ошибка при создании файла экспорта: {str(e)}")

# ======================
# УДАЛЕНИЕ ЗАДАЧ
# ======================

@dp.message_handler(commands=["deletetask"])
async def delete_task_start(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Только администратор может удалять задачи")
        return

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, task_text, status 
            FROM tasks
            ORDER BY id DESC 
            LIMIT 0
        """)
        tasks = cursor.fetchall()

        keyboard = InlineKeyboardMarkup(row_width=1)
        for task_id, task_text, status in tasks:
            keyboard.add(InlineKeyboardButton(
                f"{task_text[:30]}... (ID: {task_id}, статус: {status})", 
                callback_data=f"delete_task_{task_id}"
            ))
        
        keyboard.add(InlineKeyboardButton("✏️ Ввести ID вручную", callback_data="enter_task_id_manually_delete"))

        await bot.send_message(chat_id=message.from_user.id, text="Выберите задачу для удаления или введите ID вручную:", reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Ошибка при выборе задачи для удаления: {e}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Ошибка при получении списка задач.")

@dp.callback_query_handler(lambda c: c.data == "enter_task_id_manually_delete")
async def ask_for_manual_task_id_delete(callback_query: types.CallbackQuery):
    """Запрос ручного ввода ID задачи для удаления"""
    await bot.answer_callback_query(callback_query.id)
    await bot.send_message(chat_id=callback_query.from_user.id, text="✏️ Введите ID задачи для удаления:")
    await TaskDeletion.waiting_for_manual_id.set()

@dp.message_handler(state=TaskDeletion.waiting_for_manual_id)
async def process_manual_task_id_delete(message: types.Message, state: FSMContext):
    """Обработка ручного ввода ID задачи для удаления"""
    try:
        task_id = int(message.text)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM tasks WHERE id=?", (task_id,))
        if not cursor.fetchone():
            await bot.send_message(chat_id=message.from_user.id, text="⚠ Задача с таким ID не найдена или не принадлежит вам!")
            await state.finish()
            return
        
        await state.update_data(task_id=task_id)
        await show_delete_confirmation(message, task_id)
        await TaskDeletion.waiting_for_confirmation.set()
    except ValueError:
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Пожалуйста, введите числовой ID задачи!")
        await state.finish()
    except Exception as e:
        logger.error(f"Ошибка при обработке ручного ввода ID: {e}")
        await bot.send_message(chat_id=message.from_user.id, text="⚠ Произошла ошибка. Попробуйте снова.")
        await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith("delete_task_"))
async def select_task_for_deletion(callback_query: types.CallbackQuery):
    """Обработка выбора задачи для удаления"""
    task_id = callback_query.data.split("_")[2]
    await bot.answer_callback_query(callback_query.id)
    await show_delete_confirmation(callback_query.message, task_id)

async def show_delete_confirmation(message_obj, task_id):
    """Показать подтверждение удаления (общая функция)"""
    cursor = conn.cursor()
    cursor.execute("SELECT task_text, status, deadline FROM tasks WHERE id=?", (task_id,))
    task_info = cursor.fetchone()
    
    if not task_info:
        await bot.send_message(chat_id=message_obj.from_user.id, text="⚠ Задача не найдена!")
        return
    
    task_text, status, deadline = task_info
    
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("✅ Да, удалить", callback_data=f"confirm_deletion_{task_id}"),
        InlineKeyboardButton("❌ Нет, отменить", callback_data="cancel_deletion")
    )
    
    # Отправляем новое сообщение с подтверждением
    await bot.send_message(chat_id=message_obj.chat.id, text=
        f"Вы уверены, что хотите удалить задачу?\n\n"
        f"📌 {task_text}\n"
        f"🔄 {status}\n"
        f"⏳ {format_date(deadline) if deadline else 'нет срока'}",
        reply_markup=keyboard
    )

@dp.callback_query_handler(lambda c: c.data.startswith("confirm_deletion_"), 
                          state=TaskDeletion.waiting_for_confirmation)
async def execute_task_deletion(callback_query: types.CallbackQuery, state: FSMContext):
    """Выполнение удаления задачи"""
    try:
        task_id = callback_query.data.split("_")[2]
        
        cursor = conn.cursor()
        cursor.execute("SELECT task_text FROM tasks WHERE id=?", (task_id,))
        task = cursor.fetchone()
        
        if not task:
            await bot.send_message(chat_id=callback_query.chat.id, text="⚠ Задача не найдена!")
            return
            
        task_text = task[0]
        
        cursor.execute("DELETE FROM tasks WHERE id=?", (task_id,))
        cursor.execute("DELETE FROM tasks_log WHERE id=?", (task_id,))
        conn.commit()
        
        # Редактируем сообщение с подтверждением
        await callback_query.message.edit_text(
            f"✅ Задача успешно удалена:\n"
            f"ID: {task_id}\n"
            f"Текст: {task_text[:100]}..."
        )
        await state.finish()
    except Exception as e:
        logger.error(f"Ошибка при удалении задачи: {e}")
        await bot.send_message(chat_id=callback_query.from_user.id, text="⚠ Ошибка при удалении задачи!")

@dp.callback_query_handler(lambda c: c.data == "cancel_deletion")
async def cancel_task_deletion(callback_query: types.CallbackQuery):
    """Отмена удаления задачи"""
    await bot.answer_callback_query(callback_query.id)
    await callback_query.message.edit_text("❌ Удаление отменено.")
    await state.finish()

# ======================
# Добавление пользователя
# ======================

class AddUserState(StatesGroup):
    waiting_for_user_id = State()  # Ожидаем ID пользователя

@dp.message_handler(commands=["adduser"])
async def add_user_command(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Только администратор может добавлять пользователей")
        return
    
    # Переводим в состояние ожидания ID пользователя
    await AddUserState.waiting_for_user_id.set()
    await bot.send_message(chat_id=message.from_user.id, text="Введите ID пользователя для добавления в формате:\n'user_id|name|is_moderator|username'\n'moderator'/'username' могут быть пустыми")

@dp.message_handler(state=AddUserState.waiting_for_user_id)
async def process_user_id(message: types.Message, state: FSMContext):
    match = re.match(r'^(\d+)\|([^|]+)(?:\|(moderator))?(?:\|(.+))?$', message.text.strip())
    
    if match:
        user_id = match.group(1)
        user_name = match.group(2)
        is_moderator = match.group(3)
        username = (match.group(4) or "").strip()
    else:
        await bot.send_message(chat_id=message.from_user.id, text="Строка не соответствует формату!")
        await state.finish()
        return

    if not user_id.isdigit():
        await message.reply("ID пользователя должен быть числом!")
        await state.finish()
        return

    is_moderator = None if is_moderator == 'NULL' else is_moderator

    # Получаем подключение к базе данных из контекста
    cursor = conn.cursor()
    user_id = int(user_id)

    # Проверяем, существует ли уже пользователь
    cursor.execute("SELECT 1 FROM users WHERE tg_user_id = ?", (user_id,))
    if cursor.fetchone():
        await message.reply("⚠ Пользователь с таким ID уже существует!")
        await state.finish()
        return

    try:
        # Вставляем в базу данных
        cursor.execute('INSERT INTO users (tg_user_id, name, is_moderator, username) VALUES (?, ?, ?, ?)', (user_id, user_name, is_moderator, username))
        conn.commit()
        
        # Обновляем список разрешенных пользователей
        update_allowed_users(conn)
        update_moderator_users(conn)
        
        # Отправляем подтверждение
        await message.reply("✅ Пользователь успешно добавлен!")
        
    except sqlite3.Error as e:
        await message.reply("❌ Произошла ошибка при добавлении в базу данных")

    # Завершаем состояние после выполнения всех действий
    await state.finish()

# ======================
# Удаление Пользователя
# ======================

class RemoveUserState(StatesGroup):
    waiting_for_user_id = State()  # Ожидаем ID пользователя

@dp.message_handler(commands=["removeuser"])
async def remove_user_command(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Только администратор может удалять пользователей")
        return
    
    # Переводим в состояние ожидания ID пользователя
    await RemoveUserState.waiting_for_user_id.set()
    await bot.send_message(chat_id=message.from_user.id, text="Введите ID пользователя для удаления:")

@dp.message_handler(state=RemoveUserState.waiting_for_user_id)
async def process_remove_user(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.reply("ID пользователя должен быть числом!")
        await state.finish()
        return

    user_id = int(message.text)
    cursor = conn.cursor()
    
    # Проверяем, существует ли пользователь
    cursor.execute("SELECT 1 FROM users WHERE tg_user_id = ?", (user_id,))
    if not cursor.fetchone():
        await message.reply("⚠ Пользователь с таким ID не найден!")
        await state.finish()
        return
    
    try:
        # Удаляем пользователя из базы
        cursor.execute("DELETE FROM users WHERE tg_user_id = ?", (user_id,))
        conn.commit()
        
        # Обновляем список разрешенных пользователей
        update_allowed_users(conn)
        update_moderator_users(conn)
        
        await message.reply("✅ Пользователь успешно удален!")
        
    except sqlite3.Error as e:
        await message.reply("❌ Произошла ошибка при удалении из базы данных")
    
    await state.finish()

# ======================
# ЭКСПОРТ ПОЛЬЗОВАТЕЛЕЙ
# ======================

@dp.message_handler(commands=["export4"])
async def export_users_to_csv3(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await bot.send_message(chat_id=message.from_user.id, text="⛔ Только администратор может делать экспорт списка пользователей")
        return
      
    """Экспорт всех задач в CSV файл с кодировкой win1251"""
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT tg_user_id, name, username, is_moderator FROM users")
        users = cursor.fetchall()
        
        if not users:
            await bot.send_message(chat_id=message.from_user.id, text="📭 В базе нет пользователей.")
            return

        # Создаем CSV в памяти
        output = io.BytesIO()
        
        # Используем TextIOWrapper с нужной кодировкой
        text_buffer = io.TextIOWrapper(
            output,
            encoding='utf-8-sig',
            errors='replace',  # заменяем некодируемые символы
            newline=''
        )
        
        writer = csv.writer(
            text_buffer,
            delimiter=';',  # Указываем нужный разделитель
            quoting=csv.QUOTE_MINIMAL
        )
        
        # Заголовки столбцов
        headers = ['tg_user_id', 'name', 'username', 'is_moderator']
        writer.writerow(headers)
        
        # Данные
        for user in users:
            # Преобразуем все значения в строки
            row = [
                str(item) if item is not None else ''
                for item in user
            ]
            writer.writerow(row)
        
        # Важно: закрыть TextIOWrapper перед использованием буфера
        text_buffer.flush()
        text_buffer.detach()  # Отсоединяем TextIOWrapper от BytesIO
        output.seek(0)
        
        # Создаем временный файл
        csv_file = InputFile(output, filename="users_export.csv")
        
        await message.reply_document(
            document=csv_file
        )
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте задач: {str(e)}", exc_info=True)
        await bot.send_message(chat_id=message.from_user.id,text=f"⚠ Ошибка при создании файла экспорта: {str(e)}")

# ======================
# ID Пользователя
# ======================

@dp.message_handler(commands=["myid"])
async def get_user_id(message: types.Message):
    await bot.send_message(chat_id=message.from_user.id,text=f"Ваш 🆔 `{message.from_user.id}`", parse_mode="Markdown")

# ======================
# ФОНОВЫЕ ЗАДАЧИ
# ======================

async def check_deadlines():
    """Проверка дедлайнов и отправка напоминаний создателю"""
    while True:
        try:
            now = datetime.now().strftime("%Y-%m-%d")
            cursor = background_conn.cursor()
            cursor.execute(
                "SELECT id, chat_id, task_text, user_id, status, deadline FROM tasks "
                "WHERE deadline<=? AND status NOT IN ('исполнено','удалено')", 
                (now,)
            )
            tasks = cursor.fetchall()

            for task_id, chat_id, task_text, user_id, status, deadline in tasks:
                try:
                    # Отправляем в ЛС создателя (chat_id == user_id)
                    await bot.send_message(
                        chat_id=chat_id,
                        text=f"⏳ Напоминание о задаче 🔹{task_id}:\n📝: {task_text}\n\n👤: {user_id}\n🔄: {status} ⏳: {format_date(deadline)}"
                    )
                except exceptions.BotBlocked:
                    logger.error(f"Пользователь {chat_id} заблокировал бота")
                except exceptions.ChatNotFound:
                    logger.error(f"Чат {chat_id} не найден")
                except Exception as e:
                    logger.error(f"Ошибка: {e}")

            await asyncio.sleep(21600)
        except Exception as e:
            logger.error(f"Ошибка в фоновой задаче: {e}")
            await asyncio.sleep(60)

# ======================
# HEALTH CHECK
# ======================

async def health_check(request):
    """Endpoint для health check"""
    return web.Response(text="OK")

async def start_web_server():
    """Запуск HTTP сервера для health check"""
    app = web.Application()
    app.router.add_get('/health', health_check)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', 8000)
    await site.start()

# ======================
# ЗАПУСК БОТА
# ======================

async def main():
    """Основная функция запуска"""
    await set_bot_commands(bot)  # Регистрация команд в интерфейсе Telegram
    asyncio.create_task(check_deadlines())
    await asyncio.gather(
        start_web_server(),
        dp.start_polling()
    )

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        logger.error(f"Ошибка при запуске бота: {e}")
    finally:
        conn.close()
